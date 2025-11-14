import io
import zipfile
import base64
try:
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
except ImportError:
    openpyxl = None

from typing import List, Optional
from sayou.document.interfaces.base_parser import BaseDocumentParser
from sayou.document.models import (
    Document, Page, Sheet, TableElement, ImageElement,
    ElementMetadata, BaseElement, TableCell
)

class ExcelParser(BaseDocumentParser):
    component_name = "ExcelParser"
    SUPPORTED_TYPES = [".xlsx", ".xlsm", ".xltx", ".xltm"]

    def _parse(self, file_bytes: bytes, file_name: str, **kwargs) -> Document:
        if openpyxl is None:
            raise ImportError("openpyxl is required for ExcelParser.")

        workbook = self._load_workbook_safe(file_bytes)
        pages_list = []
        skip_hidden = kwargs.get("skip_hidden", False)
        ocr_images = kwargs.get("ocr_images", True)

        for idx, sheet_name in enumerate(workbook.sheetnames):
            sheet = workbook[sheet_name]
            is_hidden = sheet.sheet_state != "visible"
            if skip_hidden and is_hidden:
                self._log(f"Skipping hidden sheet: {sheet_name}")
                continue

            elements_list: List[BaseElement] = []

            sheet_data = [] # LLM 친화적 2D 리스트
            rows_cells = [] # High-Fidelity 2D 리스트

            for r_idx, row in enumerate(sheet.iter_rows()):
                cleaned_row_data = []
                current_row_cells = []
                
                for c_idx, cell in enumerate(row):
                    cell_text = str(cell.value) if cell.value is not None else ""
                    cleaned_row_data.append(cell_text)
                    cell_obj = TableCell(
                        text=cell_text,
                        row_span=1, 
                        col_span=1,
                        is_header=False,
                    )
                    current_row_cells.append(cell_obj)
                
                if any(cleaned_row_data):
                    sheet_data.append(cleaned_row_data)
                    rows_cells.append(current_row_cells)

            if sheet_data:
                table_elem = TableElement(
                    id=f"sheet:{idx}",
                    type="table",
                    data=sheet_data,
                    cells=rows_cells,
                    caption=f"Sheet: {sheet_name}",
                    meta=ElementMetadata(page_num=idx + 1)
                )
                elements_list.append(table_elem)
            
            sheet_images = self._extract_images(sheet, idx + 1, ocr_images)
            elements_list.extend(sheet_images)

            if elements_list:
                sheet_obj = Sheet(
                    page_num=idx + 1,
                    elements=elements_list,
                    sheet_name=sheet_name,
                    is_hidden=is_hidden,
                    sheet_index=idx
                )
                pages_list.append(sheet_obj)

        return Document(
            file_name=file_name,
            file_id=file_name,
            doc_type="sheet",
            page_count=len(pages_list),
            pages=pages_list
        )

    def _extract_images(self, sheet, page_num, ocr_enabled) -> List[ImageElement]:
        images = []
        if hasattr(sheet, "_images"):
            for i, image in enumerate(sheet._images):
                if isinstance(image, XLImage):
                    try:
                        img_elem = self._process_image_data(
                            image_bytes=image.ref.getvalue(),
                            img_format=image.format or "png",
                            elem_id=f"sheet{page_num}:img{i}",
                            page_num=page_num,
                            ocr_enabled=ocr_enabled
                        )
                        img_elem.raw_attributes["from_col"] = image.anchor._from.col
                        img_elem.raw_attributes["from_row"] = image.anchor._from.row
                        images.append(img_elem)
                    except Exception as e:
                        self._log(f"Image extraction error: {e}", level="warning")
        return images

    def _load_workbook_safe(self, file_bytes: bytes):
        """custom.xml 오류 자동 복구 로직"""
        try:
            return openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        except (TypeError, KeyError, zipfile.BadZipFile):
            self._log("⚠️ Excel load failed. Attempting repair (removing custom.xml)...")
            try:
                repaired_buffer = io.BytesIO()
                with zipfile.ZipFile(io.BytesIO(file_bytes), 'r') as zin:
                    with zipfile.ZipFile(repaired_buffer, 'w') as zout:
                        for item in zin.infolist():
                            if "custom.xml" not in item.filename:
                                buffer = zin.read(item.filename)
                                zout.writestr(item, buffer)
                repaired_buffer.seek(0)
                return openpyxl.load_workbook(repaired_buffer, data_only=True)
            except Exception as e:
                raise ValueError(f"Failed to repair Excel file: {e}")