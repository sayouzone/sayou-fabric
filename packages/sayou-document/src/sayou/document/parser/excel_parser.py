import io
import zipfile

try:
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
except ImportError:
    openpyxl = None

from typing import List

from sayou.core.registry import register_component

from ..interfaces.base_parser import BaseDocumentParser
from ..models import (
    BaseElement,
    Document,
    ElementMetadata,
    ImageElement,
    Page,
    Sheet,
    TableCell,
    TableElement,
)


@register_component("parser")
class ExcelParser(BaseDocumentParser):
    """
    (Tier 2) Parser for Microsoft Excel (.xlsx) workbooks.

    Treats each sheet as a 'Page'. Extracts cell data as tables and
    floating images embedded in the sheet. Includes a robust recovery
    mechanism for corrupted metadata XML.
    """

    component_name = "ExcelParser"
    SUPPORTED_TYPES = [".xlsx", ".xlsm", ".xltx", ".xltm"]

    @classmethod
    def can_handle(cls, file_bytes: bytes, file_name: str) -> float:
        """
        Checks for Zip signature (PK..) or OLE signature.
        """
        # XLSX (Zip)
        if file_bytes.startswith(b"PK\x03\x04") and file_name.lower().endswith(".xlsx"):
            return 1.0
        # Legacy XLS
        if file_bytes.startswith(b"\xd0\xcf\x11\xe0"):
            return 1.0
        # Extension fallback
        if any(file_name.lower().endswith(t) for t in cls.SUPPORTED_TYPES):
            return 0.8
        return 0.0

    def _do_parse(self, file_bytes: bytes, file_name: str, **kwargs) -> Document:
        """
        Parse Excel bytes into a structured Document.

        Args:
            file_bytes (bytes): Binary content of the .xlsx file.
            file_name (str): Original filename.
            **kwargs:
                - skip_hidden (bool): If True, ignore hidden sheets.
                - ocr_images (bool): If True, run OCR on embedded images.

        Returns:
            Document: A document object with 'doc_type="sheet"'.
        """
        if openpyxl is None:
            raise ImportError("openpyxl is required for ExcelParser.")

        workbook = self._load_workbook_safe(file_bytes)
        pages_list = []
        skip_hidden = kwargs.get("skip_hidden", False)
        ocr_images = kwargs.get("ocr_images", True)

        for idx, sheet_name in enumerate(workbook.sheetnames):
            sheet = workbook[sheet_name]
            is_hidden = sheet.sheet_state != "visible"
            if skip_hidden and is_hidden:
                self._log(f"Skipping hidden sheet: {sheet_name}")
                continue

            elements_list: List[BaseElement] = []

            sheet_data = []
            rows_cells = []

            for r_idx, row in enumerate(sheet.iter_rows()):
                cleaned_row_data = []
                current_row_cells = []

                for c_idx, cell in enumerate(row):
                    cell_text = str(cell.value) if cell.value is not None else ""
                    cleaned_row_data.append(cell_text)
                    cell_obj = TableCell(
                        text=cell_text,
                        row_span=1,
                        col_span=1,
                        is_header=False,
                    )
                    current_row_cells.append(cell_obj)

                if any(cleaned_row_data):
                    sheet_data.append(cleaned_row_data)
                    rows_cells.append(current_row_cells)

            if sheet_data:
                table_elem = TableElement(
                    id=f"sheet:{idx}",
                    type="table",
                    data=sheet_data,
                    cells=rows_cells,
                    caption=f"Sheet: {sheet_name}",
                    meta=ElementMetadata(page_num=idx + 1),
                )
                elements_list.append(table_elem)

            sheet_images = self._extract_images(sheet, idx + 1, ocr_images)
            elements_list.extend(sheet_images)

            if elements_list:
                sheet_obj = Sheet(
                    page_num=idx + 1,
                    elements=elements_list,
                    sheet_name=sheet_name,
                    is_hidden=is_hidden,
                    sheet_index=idx,
                )
                pages_list.append(sheet_obj)

        return Document(
            file_name=file_name,
            file_id=file_name,
            doc_type="sheet",
            page_count=len(pages_list),
            pages=pages_list,
        )

    def _extract_images(self, sheet, page_num, ocr_enabled) -> List[ImageElement]:
        """
        Extract embedded images (drawings) from a worksheet.

        Args:
            sheet (Worksheet): The openpyxl worksheet object.
            page_num (int): The sheet index (1-based).
            ocr_enabled (bool): Whether to perform OCR.

        Returns:
            List[ImageElement]: Extracted images with position metadata (row/col).
        """
        images = []
        if hasattr(sheet, "_images"):
            for i, image in enumerate(sheet._images):
                if isinstance(image, XLImage):
                    try:
                        img_elem = self._process_image_data(
                            image_bytes=image.ref.getvalue(),
                            img_format=image.format or "png",
                            elem_id=f"sheet{page_num}:img{i}",
                            page_num=page_num,
                            ocr_enabled=ocr_enabled,
                        )
                        img_elem.raw_attributes["from_col"] = image.anchor._from.col
                        img_elem.raw_attributes["from_row"] = image.anchor._from.row
                        images.append(img_elem)
                    except Exception as e:
                        self._log(f"Image extraction error: {e}", level="warning")
        return images

    def _load_workbook_safe(self, file_bytes: bytes):
        """
        Load workbook with fallback repair logic.

        Attempts to load normally. If it fails due to common XML errors
        (e.g., 'custom.xml' issues from 3rd party tools), it rewrites
        the ZIP structure to exclude the problematic file and retries.

        Args:
            file_bytes (bytes): The raw file content.

        Returns:
            Workbook: The loaded openpyxl Workbook object.
        """
        try:
            return openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        except (TypeError, KeyError, zipfile.BadZipFile):
            self._log("⚠️ Excel load failed. Attempting repair (removing custom.xml)...")
            try:
                repaired_buffer = io.BytesIO()
                with zipfile.ZipFile(io.BytesIO(file_bytes), "r") as zin:
                    with zipfile.ZipFile(repaired_buffer, "w") as zout:
                        for item in zin.infolist():
                            if "custom.xml" not in item.filename:
                                buffer = zin.read(item.filename)
                                zout.writestr(item, buffer)
                repaired_buffer.seek(0)
                return openpyxl.load_workbook(repaired_buffer, data_only=True)
            except Exception as e:
                raise ValueError(f"Failed to repair Excel file: {e}")
